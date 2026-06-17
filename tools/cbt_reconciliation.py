import streamlit as st
import google.generativeai as genai
import base64
import json
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re

def ask_ollama(prompt, model="qwen2.5:7b"):

    response = requests.post(
        "http://localhost:11434/api/generate",
        json={
            "model": model,
            "prompt": prompt,
            "stream": False
        }
    )

    return response.json()["response"]


def clean_text_with_ollama(text):

    prompt = f"""
Clean and normalize this OCR extracted text.

Rules:
- Keep original meaning
- Do NOT invent values
- Do NOT guess unclear handwritten numbers
- Preserve registration numbers carefully
- Remove OCR noise
- Fix spacing issues
- Standardize capitalization
- Return ONLY cleaned text

TEXT:
{text}
"""

    try:
        response = ask_ollama(prompt)

        return response.strip()

    except:
        return text

st.set_page_config(page_title="Attendance Report", layout="wide")
st.title("CBT – Attendance Reconciliation Tool")
st.caption("Akshay Singh | iTest Content Team, SIFY Technologies")

api_key = st.text_input(
    "Enter your Gemini API Key",
    type="password",
    placeholder="AIzaSy..."
)
col1, col2 = st.columns(2)
with col1:
    biometric_file = st.file_uploader("Upload Biometric PDF (Form C11)", type="pdf")
with col2:
    attendance_file = st.file_uploader("Upload Attendance PDF", type="pdf")


def pdf_to_base64(uploaded_file):
    uploaded_file.seek(0)
    return base64.standard_b64encode(uploaded_file.read()).decode("utf-8")


def extract_json_safe(text):
    text = re.sub(r'```(?:json)?', '', text).strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    start = text.find('[')
    if start == -1:
        raise ValueError(f"No JSON array found. Response snippet:\n{text[:500]}")
    depth = 0
    end = -1
    in_string = False
    escape_next = False
    for i, ch in enumerate(text[start:], start):
        if escape_next:
            escape_next = False
            continue
        if ch == '\\' and in_string:
            escape_next = True
            continue
        if ch == '"' and not escape_next:
            in_string = not in_string
            continue
        if not in_string:
            if ch == '[':
                depth += 1
            elif ch == ']':
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break
    if end == -1:
        partial = text[start:]
        partial += '}' * max(0, partial.count('{') - partial.count('}'))
        partial += ']' * max(0, partial.count('[') - partial.count(']'))
        try:
            return json.loads(partial)
        except Exception:
            raise ValueError(f"Could not parse JSON. Snippet:\n{text[start:start+400]}")
    try:
        return json.loads(text[start:end])
    except Exception as e:
        raise ValueError(f"JSON parse error: {e}")

def call_gemini(model, prompt, pdf_bytes, label):
    for attempt in range(3):
        try:
            response = model.generate_content(
                contents=[
                    {
                        "mime_type": "application/pdf",
                        "data": pdf_bytes
                    },
                    prompt
                ],
                generation_config={
                    "temperature": 0,
                    "response_mime_type": "application/json"
                }
            )

            text = response.text
            records = extract_json_safe(text)

            return records

        except Exception as e:
            if attempt < 2:
                st.warning(f"Attempt {attempt+1} failed for {label}, retrying... ({e})")
                continue
            raise e

    return []

def extract_biometric(client, b64_pdf):
    prompt = """You are extracting data from an IIBF CBT Biometric Resource Tracker (Form C11).
The PDF has multiple pages. Each page has a table with columns:
Sl. No. | Membership No/Registration No | Candidate Name | Photo Captured (Yes/No) | Lab No Assigned | Seat Number Assigned

Handwritten entries are present. Extract ALL rows from ALL pages without skipping any.

IMPORTANT: Return ONLY a valid JSON array. No explanation, no markdown. Start with [ end with ].

[
  {
    "sl_no_bio": "1",
    "membership_no": "500070716",
    "candidate_name": "Suresh Kumar",
    "photo_captured": "Y",
    "lab_no": "L1",
    "seat_number": "L1-40",
    "remark": ""
  }
]

Rules:
- sl_no_bio: serial number from Sl. No. column — extract exactly as written
- seat_number: copy exactly as written e.g. "L1-40" or "L2-5"
- If any field is unreadable, set remark: "Manual verification required – [reason]"
- Skip only completely blank rows
- No text outside the JSON array
"""
    
    return call_gemini(
    client,
    prompt,
    base64.b64decode(b64_pdf),
    "Biometric PDF"
)


def extract_attendance(client, b64_pdf):
    prompt = """You are extracting data from an IIBF exam Attendance Sheet PDF.
Multiple pages. Each page has a table:
Sl. No. | Ticket No. | Membership No/Candidate Name/DOB/Category | Photo and Signature | Signature | Signature After Bio Break | Left Hand Thumb Impression (LTI) | System No.

Third column example: "500070716 / Mr. SURESH KUMAR A / 26-12-1983 / -"

Extract ALL rows from ALL pages without skipping any.

IMPORTANT: Return ONLY a valid JSON array. No explanation, no markdown. Start with [ end with ].

[
  {
    "sl_no_att": "1",
    "ticket_no": "154000",
    "membership_no": "500070716",
    "candidate_name": "SURESH KUMAR A",
    "signature_present": "YES",
    "system_no": "40",
    "remark": ""
  }
]

Rules:
- sl_no_att: serial number from Sl. No. column — extract exactly as written
- signature_present: "ABSENT" = stamped ABSENT | "YES" = signature visible | "NO" = blank but not ABSENT
- system_no: the value in the rightmost column labelled "SYSTEM NO." — physically the last column on the far right edge of the table. Handwritten number, usually 1-2 digits. IMPORTANT handwriting notes: the number 7 in this document is written with a horizontal cross-stroke through the middle, making it look similar to 4 or f — if you see a vertical stroke with a horizontal cross and a diagonal, it is 7 not 4. The number 1 may look like a slash. Sometimes written in two lines like "L1" above and "17" below — combine as "L1-17". Do NOT confuse with Sl. No. (leftmost column) or Ticket No. (second column). If empty write "". If genuinely unclear after careful reading, write "Manual verification required"
- membership_no: numeric part only e.g. "500070716"
- candidate_name: name only, remove Mr./Ms./Mrs. prefix
- If any field is unreadable, set remark: "Manual verification required – [reason]"
- No text outside the JSON array
"""
    
    return call_gemini(
    client,
    prompt,
    base64.b64decode(b64_pdf),
    "Attendance PDF"
)


def names_match(name1, name2):
    if not name1 or not name2:
        return False
    n1 = re.sub(r'[^a-z ]', '', name1.lower().strip())
    n2 = re.sub(r'[^a-z ]', '', name2.lower().strip())
    w1, w2 = set(n1.split()), set(n2.split())
    if not w1 or not w2:
        return False
    return len(w1 & w2) / max(len(w1), len(w2)) >= 0.5


def seat_system_match(seat_no, system_no):
    """
    Extract numeric part from seat_no (e.g. 'L1-40' -> '40')
    and compare with system_no (e.g. '40').
    Returns: 'MATCH', 'MISMATCH', or 'MISSING'
    """
    if not seat_no or str(seat_no).strip() in ("—", "", "None"):
        return "MISSING"
    if not system_no or str(system_no).strip() in ("—", "", "None"):
        return "MISSING"

    seat_clean   = str(seat_no).strip()
    system_clean = str(system_no).strip()

    # Extract number after last hyphen in seat_no
    parts = seat_clean.split("-")
    seat_num = parts[-1].strip() if len(parts) >= 2 else seat_clean

    try:
        return "MATCH" if int(seat_num) == int(system_clean) else "MISMATCH"
    except ValueError:
        return "MATCH" if seat_num.lower() == system_clean.lower() else "MISMATCH"


def determine_match_status(name_ok, seat_sys_status, membership_found):
    """
    MATCHED = membership found + name OK + seat/system MATCH
    REVIEW  = membership found but any check fails or value missing
    MISSING = membership not found in one PDF
    """
    if not membership_found:
        return "MISSING"
    if not name_ok:
        return "REVIEW"
    if seat_sys_status in ("MISMATCH", "MISSING"):
        return "REVIEW"
    return "MATCHED"


def build_report(biometric_data, attendance_data):

    def normalize_mno(raw):
        s = re.sub(r'\s+', '', str(raw).strip())
        s = s.lstrip('-')
        s = re.sub(r'[,./\\]+$', '', s)
        s = re.sub(r'[-_,.]([D0O])([1lI)]*)\d*[,./]*$', '', s, flags=re.IGNORECASE)
        s = re.sub(r'(D[iIlL1)]+)\d*[,./]*$', '', s, flags=re.IGNORECASE)
        s = re.sub(r'[,./\\]+$', '', s)
        return s

    att_lookup = {}
    for row in attendance_data:
        mno = normalize_mno(str(row.get("membership_no", "")))
        if mno:
            att_lookup[mno] = row

    report_rows = []


    for bio_row in biometric_data:
        mno      = normalize_mno(bio_row.get("membership_no", ""))
        bio_name = bio_row.get("candidate_name", "")
        seat_no  = bio_row.get("seat_number", "")
        bio_rem  = bio_row.get("remark", "")
        sl_bio   = bio_row.get("sl_no_bio", "")
        att_row  = att_lookup.get(mno)

        if att_row:
            att_name  = att_row.get("candidate_name", "")
            signature = att_row.get("signature_present", "")
            system_no = att_row.get("system_no", "")
            att_rem   = att_row.get("remark", "")
            sl_att    = att_row.get("sl_no_att", "")
            mno_att   = att_row.get("membership_no", "")

            name_ok         = names_match(bio_name, att_name)
            seat_sys_status = seat_system_match(seat_no, system_no)
            match_status    = determine_match_status(name_ok, seat_sys_status, True)
            mno_verify      = "MATCH" if str(mno).strip() == str(mno_att).strip() else "MISMATCH"

            remarks = []
            if bio_rem: remarks.append(f"Biometric: {bio_rem}")
            if att_rem: remarks.append(f"Attendance: {att_rem}")
            if not name_ok:
                remarks.append(f"Name mismatch – Bio: '{bio_name}' vs Att: '{att_name}'")
            if seat_sys_status == "MISMATCH":
                remarks.append(f"Seat/System mismatch – Seat: '{seat_no}' vs System No: '{system_no}'")
            if seat_sys_status == "MISSING":
                remarks.append(f"Seat/System No missing – Seat: '{seat_no}', System No: '{system_no}'")

            report_rows.append({
                "Sl No (Biometric)":           sl_bio,
                "Membership No (Biometric)":   mno,
                "Sl No (Attendance)":          sl_att,
                "Membership No (Attendance)":  mno_att,
                "Membership Verification":     mno_verify,
                "Candidate Name (Biometric)":  bio_name,
                "Candidate Name (Attendance)": att_name,
                "Name Verification":           "MATCH" if name_ok else "MISMATCH",
                "Seat No (Biometric PDF)":     seat_no,
                "System No (Attendance PDF)":  system_no,
                "Seat vs System Verification": seat_sys_status,
                "Signature (Attendance PDF)":  signature,
                "Final Match Status":          match_status,
                "Remarks":                     " | ".join(remarks) if remarks else ""
            })
        else:
            remarks = ["Membership No not found in Attendance PDF"]
            if bio_rem: remarks.append(f"Biometric: {bio_rem}")
            report_rows.append({
                "Sl No (Biometric)":           sl_bio,
                "Membership No (Biometric)":   mno,
                "Sl No (Attendance)":          "—",
                "Membership No (Attendance)":  "NOT FOUND",
                "Membership Verification":     "NOT FOUND",
                "Candidate Name (Biometric)":  bio_name,
                "Candidate Name (Attendance)": "NOT FOUND",
                "Name Verification":           "NOT FOUND",
                "Seat No (Biometric PDF)":     seat_no,
                "System No (Attendance PDF)":  "—",
                "Seat vs System Verification": "NOT FOUND",
                "Signature (Attendance PDF)":  "—",
                "Final Match Status":          "MISSING",
                "Remarks":                     " | ".join(remarks)
            })

    bio_mnos = {str(r.get("membership_no", "")).strip() for r in biometric_data}
    for mno, att_row in att_lookup.items():
        if mno not in bio_mnos:
            remarks = ["Membership No not found in Biometric PDF"]
            if att_row.get("remark"): remarks.append(f"Attendance: {att_row['remark']}")
            report_rows.append({
                "Sl No (Biometric)":           "—",
                "Membership No (Biometric)":   "NOT FOUND",
                "Sl No (Attendance)":          att_row.get("sl_no_att", ""),
                "Membership No (Attendance)":  mno,
                "Membership Verification":     "NOT FOUND",
                "Candidate Name (Biometric)":  "NOT FOUND",
                "Candidate Name (Attendance)": att_row.get("candidate_name", ""),
                "Name Verification":           "NOT FOUND",
                "Seat No (Biometric PDF)":     "—",
                "System No (Attendance PDF)":  att_row.get("system_no", ""),
                "Seat vs System Verification": "NOT FOUND",
                "Signature (Attendance PDF)":  att_row.get("signature_present", ""),
                "Final Match Status":          "MISSING",
                "Remarks":                     " | ".join(remarks)
            })

    return report_rows


def generate_excel(report_rows, bio_total, att_total):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation Report"

    hdr_fill      = PatternFill("solid", fgColor="1F4E79")
    matched_fill  = PatternFill("solid", fgColor="C6EFCE")
    matched_alt   = PatternFill("solid", fgColor="EAF7EE")
    review_fill   = PatternFill("solid", fgColor="FFEB9C")
    missing_fill  = PatternFill("solid", fgColor="FFC7CE")
    green_fill    = PatternFill("solid", fgColor="00B050")
    red_fill      = PatternFill("solid", fgColor="FF0000")
    amber_fill    = PatternFill("solid", fgColor="FFC000")

    thin   = Side(style='thin', color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "CBT – Attendance Reconciliation Report"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"] = "Akshay Singh  |  iTest Content Team  |  SIFY Technologies"
    ws["A2"].font      = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "Sl No\n(Biometric)",
        "Membership No\n(Biometric)",
        "Sl No\n(Attendance)",
        "Membership No\n(Attendance)",
        "Membership\nVerification",
        "Candidate Name\n(Biometric)",
        "Candidate Name\n(Attendance)",
        "Name\nVerification",
        "Seat No\n(Biometric PDF)",
        "System No\n(Attendance PDF)",
        "Seat vs System\nVerification",
        "Signature\n(Attendance PDF)",
        "Final Match\nStatus",
        "Remarks"
    ]
    col_widths = [10, 20, 10, 20, 16, 26, 26, 14, 16, 16, 18, 14, 14, 55]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 40

    col_keys = [
        "Sl No (Biometric)",
        "Membership No (Biometric)",
        "Sl No (Attendance)",
        "Membership No (Attendance)",
        "Membership Verification",
        "Candidate Name (Biometric)",
        "Candidate Name (Attendance)",
        "Name Verification",
        "Seat No (Biometric PDF)",
        "System No (Attendance PDF)",
        "Seat vs System Verification",
        "Signature (Attendance PDF)",
        "Final Match Status",
        "Remarks"
    ]

    for ri, row in enumerate(report_rows, 4):
        status    = row.get("Final Match Status", "")
        base_fill = (matched_fill if ri % 2 == 0 else matched_alt) \
                    if status == "MATCHED" \
                    else review_fill if status == "REVIEW" \
                    else missing_fill

        for ci, key in enumerate(col_keys, 1):
            val  = row.get(key, "")

            # ── Excel formulas for auto-updating verification columns ──
            if ci == 5:    # Membership Verification
                write_val = f'=IF(D{ri}="NOT FOUND","NOT FOUND",IF(B{ri}=D{ri},"MATCH","MISMATCH"))'

            elif ci == 8:  # Name Verification
                write_val = (f'=IF(OR(F{ri}="NOT FOUND",G{ri}="NOT FOUND"),"NOT FOUND",'
                             f'IF(TRIM(UPPER(F{ri}))=TRIM(UPPER(G{ri})),"MATCH","MISMATCH"))')

            elif ci == 11: # Seat vs System Verification
                write_val = (f'=IF(OR(I{ri}="—",I{ri}="",J{ri}="—",J{ri}=""),"MISSING",'
                             f'IFERROR(IF(VALUE(MID(I{ri},FIND("-",I{ri})+1,LEN(I{ri})))=VALUE(J{ri}),"MATCH","MISMATCH"),"MISSING"))')

            elif ci == 13: # Final Match Status
                write_val = (f'=IF(E{ri}="NOT FOUND","MISSING",'
                             f'IF(OR(H{ri}="MISMATCH",K{ri}="MISMATCH",K{ri}="MISSING"),"REVIEW",'
                             f'IF(AND(E{ri}="MATCH",H{ri}="MATCH",K{ri}="MATCH"),"MATCHED","REVIEW")))')

            else:
                write_val = val

            cell = ws.cell(row=ri, column=ci, value=write_val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill      = base_fill

            if ci in (2, 4):
                cell.font = Font(bold=True)

            # Membership Verification — col 5
            if ci == 5:
                if val == "MATCH":
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "MISMATCH":
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "NOT FOUND":
                    cell.fill = amber_fill
                    cell.font = Font(bold=True)

            # Name Verification — col 8
            elif ci == 8:
                if val == "MATCH":
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "MISMATCH":
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "NOT FOUND":
                    cell.fill = amber_fill
                    cell.font = Font(bold=True)

            # Seat vs System Verification — col 11
            elif ci == 11:
                if val == "MATCH":
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "MISMATCH":
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "MISSING":
                    cell.fill = amber_fill
                    cell.font = Font(bold=True)
                elif val == "NOT FOUND":
                    cell.fill = amber_fill
                    cell.font = Font(bold=True)

            # Final Match Status — col 13
            elif ci == 13:
                if val == "MATCHED":
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val == "REVIEW":
                    cell.fill = review_fill
                    cell.font = Font(bold=True)
                elif val == "MISSING":
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color="FFFFFF")

        ws.row_dimensions[ri].height = 18

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 48
    ws2.column_dimensions["B"].width = 15

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Reconciliation Summary"
    ws2["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill      = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 24

    total     = len(report_rows)
    matched   = sum(1 for r in report_rows if r["Final Match Status"] == "MATCHED")
    review    = sum(1 for r in report_rows if r["Final Match Status"] == "REVIEW")
    missing   = sum(1 for r in report_rows if r["Final Match Status"] == "MISSING")
    absent    = sum(1 for r in report_rows if r.get("Signature (Attendance PDF)") == "ABSENT")
    sig_no    = sum(1 for r in report_rows if r.get("Signature (Attendance PDF)") == "NO")
    seat_mis  = sum(1 for r in report_rows if r.get("Seat vs System Verification") == "MISMATCH")
    seat_miss = sum(1 for r in report_rows if r.get("Seat vs System Verification") == "MISSING")
    name_mis  = sum(1 for r in report_rows if r.get("Name Verification") == "MISMATCH")
    mno_mis   = sum(1 for r in report_rows if r.get("Membership Verification") == "MISMATCH")

    sec_fill  = PatternFill("solid", fgColor="D6E4F0")
    even_fill = PatternFill("solid", fgColor="F2F2F2")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")

    summary_rows = [
        ("── PDF COUNTS ──",                               None),
        ("Total Candidates in Biometric PDF",              bio_total),
        ("Total Candidates in Attendance PDF",             att_total),
        ("Total Records in Reconciliation Report",         total),
        ("",                                               None),
        ("── FINAL MATCH STATUS ──",                       None),
        ("✅  Fully Matched (All checks passed)",           matched),
        ("⚠️  Review Required (At least one check failed)", review),
        ("❌  Missing (Not found in one PDF)",              missing),
        ("",                                               None),
        ("── INDIVIDUAL CHECK RESULTS ──",                 None),
        ("Membership No Mismatch count",                   mno_mis),
        ("Name Mismatch count",                            name_mis),
        ("Seat No vs System No — MISMATCH count",          seat_mis),
        ("Seat No vs System No — MISSING value count",     seat_miss),
        ("",                                               None),
        ("── ATTENDANCE STATUS ──",                        None),
        ("Marked ABSENT in Attendance PDF",                absent),
        ("Signature Not Present (blank, not absent)",      sig_no),
    ]

    for i, (label, val) in enumerate(summary_rows, 2):
        ca = ws2.cell(row=i, column=1, value=label)
        cb = ws2.cell(row=i, column=2, value=val)
        ws2.row_dimensions[i].height = 18
        if val is None:
            ca.font = Font(bold=True, size=10)
            ca.fill = sec_fill
            cb.fill = sec_fill
        else:
            fill = even_fill if i % 2 == 0 else odd_fill
            ca.font = Font(size=10)
            ca.fill = fill
            cb.font = Font(bold=True, size=10)
            cb.fill = fill
            cb.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Generate Report button ────────────────────────────────────────────────────
if st.button(
    "Generate Report",
    type="primary",
    disabled=not (api_key and biometric_file and attendance_file)
):

    st.session_state["report_rows"] = []
    st.session_state["excel_bytes"] = None
    st.session_state["bio_total"] = 0
    st.session_state["att_total"] = 0

    with st.spinner("Step 1/4: Reading Biometric PDF with AI..."):
        try:
            genai.configure(api_key=api_key)

            client = genai.GenerativeModel(
                model_name="gemini-2.5-flash"
            )

            biometric_data = extract_biometric(
                client,
                pdf_to_base64(biometric_file)
            )

            st.session_state["bio_total"] = len(biometric_data)

            st.success(
                f"✅ Biometric PDF: {len(biometric_data)} records extracted"
            )

        except Exception as e:
            st.error(f"Error reading Biometric PDF: {e}")
            st.stop()

    with st.spinner(
        "Step 2/4: Reading Attendance PDF with AI (large file — 3–5 min)..."
    ):
        try:
            attendance_data = extract_attendance(
                client,
                pdf_to_base64(attendance_file)
            )

            st.session_state["att_total"] = len(attendance_data)

            st.success(
                f"✅ Attendance PDF: {len(attendance_data)} records extracted"
            )

        except Exception as e:
            st.error(f"Error reading Attendance PDF: {e}")
            st.stop()

    with st.spinner("Step 3/4: Cross-verifying records..."):
        report_rows = build_report(
            biometric_data,
            attendance_data
        )

        st.session_state["report_rows"] = report_rows

        st.success(
            f"✅ Reconciliation complete: {len(report_rows)} total records"
        )

    with st.spinner("Step 4/4: Generating Excel..."):
        excel_buf = generate_excel(
            report_rows,
            bio_total=st.session_state["bio_total"],
            att_total=st.session_state["att_total"]
        )

        st.session_state["excel_bytes"] = excel_buf.getvalue()

        st.success("✅ Excel ready — click Download below!")

# ── Results — persistent via session_state ────────────────────────────────────
if st.session_state.get("report_rows"):
    report_rows = st.session_state["report_rows"]

    matched_n  = sum(1 for r in report_rows if r["Final Match Status"] == "MATCHED")
    review_n   = sum(1 for r in report_rows if r["Final Match Status"] == "REVIEW")
    missing_n  = sum(1 for r in report_rows if r["Final Match Status"] == "MISSING")
    seat_mis_n = sum(1 for r in report_rows if r.get("Seat vs System Verification") == "MISMATCH")

    st.divider()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("✅ Fully Matched",        matched_n)
    c2.metric("⚠️ Review Required",      review_n)
    c3.metric("❌ Missing",              missing_n)
    c4.metric("🔴 Seat/System Mismatch", seat_mis_n)

    st.download_button(
        label="⬇️ Download Excel Report",
        data=st.session_state["excel_bytes"],
        file_name="Attendance_Reconciliation_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )

    issues = [r for r in report_rows if r["Final Match Status"] in ("REVIEW", "MISSING")]
    if issues:
        st.subheader("⚠️ Items Requiring Manual Verification")
        st.dataframe(pd.DataFrame(issues), use_container_width=True)

    with st.expander("📋 View Full Report"):
        st.dataframe(pd.DataFrame(report_rows), use_container_width=True)
